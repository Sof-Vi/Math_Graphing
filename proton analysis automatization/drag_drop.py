# Python 3.13
# Code was made using gpt

import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT_SUFFIX = "_organized.xlsx"


# -------------------------
# AUTO DETECTION HELPERS
# -------------------------

def detect_mode(line_sample: str):
    """
    Try to detect if file is FRAM or FLASH based on content.
    """
    if "expected_data" in line_sample or "Read_Data" in line_sample:
        return "FRAM"
    return "FLASH"


# -------------------------
# FRAM PARSER 
# -------------------------

def parse_fram(line):

    row = ["", "", "", "", "", 0]

    dt = re.search(r"\d{4}-\d{2}-\d{2}\s+\d+:\d+:\d+\.\d+", line)
    if dt:
        row[0] = dt.group()

    event = re.search(r"\[[^\]]+\]", line)
    if event:
        row[1] = event.group()

    mev = re.search(r"\b0x[0-9A-Fa-f]{3,4}\b", line)
    if mev:
        row[2] = mev.group()

    expected = re.search(r"expected_data\s*=\s*([^\s,]+)", line, re.IGNORECASE)
    if expected:
        row[3] = expected.group(1)

    read = re.search(r"Read_Data\s*=\s*([^\s,]+)", line, re.IGNORECASE)
    if read:
        row[4] = read.group(1)

    err = re.search(r"err\s*=\s*(\d+)", line, re.IGNORECASE)
    if err:
        row[5] = 1 if int(err.group(1)) > 0 else 0

    return row


# -------------------------
# FLASH PARSER 
# -------------------------

def parse_flash(line):

    row = ["", "", "", "", 0]

    dt = re.search(r"(\d{4}-\d{2}-\d{2})\s+(\d+:\d+:\d+\.\d+)", line)
    if dt:
        row[0] = dt.group()

    event = re.search(r"\[[^\]]+\]", line)
    if event:
        row[1] = event.group()

    # MEV up to 6–7 hex chars
    mev = re.search(r"\b0x[0-9A-Fa-f]{3,7}\b", line)
    if mev:
        row[2] = mev.group()

    err = re.search(r"err\s*=\s*(\d+)", line, re.IGNORECASE)
    if err:
        row[3] = int(err.group(1))

    return row


# -------------------------
# AUTOFIT
# -------------------------

def autofit(ws):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
            except:
                pass

        ws.column_dimensions[letter].width = min(max_len + 4, 60)


# -------------------------
# MAIN EXCEL CREATION
# -------------------------

def create_excel(txt_path):

    txt_path = Path(txt_path).expanduser().resolve()

    if not txt_path.exists():
        raise FileNotFoundError(f"File not found: {txt_path}")

    wb = Workbook()
    ws = wb.active

    ws.append(["Date & Time", "Event", "MEV", "Expected/Data Read or Err", "Extra"])

    with open(txt_path, "r", encoding="utf-8", errors="ignore") as f:

        lines = [l.strip() for l in f if l.strip()]

        if not lines:
            print("Empty file")
            return

        mode = detect_mode(lines[0])
        print(f"Detected mode: {mode}")

        for line in lines:

            if mode == "FRAM":
                ws.append(parse_fram(line))
            else:
                ws.append(parse_flash(line))

    ws.freeze_panes = "A2"

    last_row = ws.max_row
    last_col = ws.max_column

    table = Table(
        displayName="LOG_TABLE",
        ref=f"A1:{get_column_letter(last_col)}{last_row}"
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    autofit(ws)

    output_path = txt_path.parent / f"{txt_path.stem}{OUTPUT_SUFFIX}"

    wb.save(output_path)

    print("\nDone!")
    print(output_path)


# -------------------------
# CLI
# -------------------------

if __name__ == "__main__":

    if len(sys.argv) < 2:
        print("\nUsage:")
        print("python drag_drop.py path/to/file.txt")
    else:
        create_excel(sys.argv[1])